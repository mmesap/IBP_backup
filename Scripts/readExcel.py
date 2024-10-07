import openpyxl

name = "Softtek"

book = openpyxl.load_workbook("./Clean/"+name + 'ForClean.xlsx')
sheet = book.active
for column in sheet.columns:
    # Elimina los "No registra"
    for cell in column:
        if cell.value == "No registra" or cell.value == "No Registra" or cell.value == "null" or cell.value == "Null" or cell.value == "No tiene":
            cell.value = None
    
    if column[0].value == "NOMBRE COMERCIAL EMPRESA" or column[0].value == "PAÍS":     
        for cell in column:
            if cell.value is not None :
                # Mayúscula y quitar espacios
                cell.value = str(cell.value).upper().strip()
                cell.value = str(cell.value).replace("  ", " ").replace(",", "").replace(".", "").replace("(", "").replace(")", "")

    if column[0].value == "TELÉFONO DE EMPRESA" or column[0].value == "TELÉFONO EMPRESA" or column[0].value == "TELEFONO 1" or column[0].value == "TELEFONO 2" or column[0].value == "CELULAR":
        for cell in column:
            if cell.value is not None :
                # Quitar + y espacios
                cell.value = str(cell.value).replace("+", "").strip().replace(" ", "").replace(".", "").replace("(", "").replace(")", "").replace("-", "").replace(",", "/").replace("PBX:", "").replace("Y", "/").replace("y", "/")
            
                if len(str(cell.value).split("/")) > 1:
                    if sheet.cell(row=1, column=cell.column + 1).value != str(column[0].value + " OTRO") :
                        sheet.insert_cols(cell.column + 1, 1)
                    sheet.cell(row=1, column=cell.column + 1).value = str(column[0].value + " OTRO") 
                    sheet.cell(row=cell.row, column=cell.column + 1).value = str(cell.value).split("/")[1]
                    sheet.cell(row=cell.row, column=cell.column ).value = str(cell.value).split("/")[0]
                cell.value = "+" + str(cell.value).split("/")[0]

    if column[0].value == "INTERLOCUTOR 1":
        for cell in column:
            list_full_name = str(cell.value).strip().split(" ")
            if len(list_full_name) > 2:
                cell.value = list_full_name[0] + " " + list_full_name[1]
                sheet.cell(row=cell.row, column=column[0].column + 1).value = " ".join(list_full_name[2:])
            elif len(list_full_name) == 2: 
                cell.value = list_full_name[0]
                sheet.cell(row=cell.row, column=column[0].column + 1).value = list_full_name[1]
        column[0].value = "NOMBRES"
        sheet.cell(row=1, column=column[0].column + 1).value = "APELLIDOS"
        sheet.cell(row=1, column=column[0].column + 2).value = "OPORTUNIDAD"

    if column[0].value == "OPORTUNIDAD":
        for cell in column:
            cell.value = format('{} / MQL / {} {}').format(sheet.cell(row=cell.row, column=column[0].column - 15).value, sheet.cell(row=cell.row, column=column[0].column - 2).value, sheet.cell(row=cell.row, column=column[0].column - 1).value) 
        column[0].value = "OPORTUNIDAD"
book.save("./Clientify/"+name +'ForClientify.xlsx')



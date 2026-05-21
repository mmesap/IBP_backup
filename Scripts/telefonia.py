import os
import re
import pandas as pd
import unicodedata
from openpyxl import load_workbook

# ==============================
# RUTAS
# ==============================
INPUT_XLSX = "/Users/elizabethmesa/ibp/Telefonia/BASE.xlsx"
OUTPUT_DIR = "/Users/elizabethmesa/ibp/Telefonia"

# ==============================
# COLUMNAS DE SALIDA
# ==============================
OUTPUT_COLUMNS = [
    "name", "phone", "email", "email2", "title", "company",
    "address", "city", "state", "zip", "country", "usc",
    "website", "campana"
]

# ==============================
# UTILIDADES DE LIMPIEZA
# ==============================
def safe_str(value):
    if pd.isna(value) or value is None:
        return ""
    s = str(value).strip()
    return "" if s.lower() in ("nan", "none") else s

def strip_accents(s: str) -> str:
    if not s:
        return ""
    return "".join(
        ch for ch in unicodedata.normalize("NFKD", s)
        if not unicodedata.combining(ch)
    )

def clean_cell(value) -> str:
    """
    Limpieza general (NO usar para PAÍS):
    - NaN → ""
    - quitar tildes
    - quitar comas
    - quitar saltos de línea
    """
    s = safe_str(value)
    s = strip_accents(s)
    s = s.replace(",", "")
    s = s.replace("\r", " ").replace("\n", " ").strip()
    return s

def sanitize_filename(name: str) -> str:
    name = clean_cell(name)
    name = re.sub(r'[\\/:*?"<>|]+', "_", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name if name else "campana"

# ==============================
# TELÉFONOS
# ==============================
import re

def normalize_phone_piece(value) -> str:
    """
    Limpia el teléfono sin romper '+'.
    Quita espacios, guiones, paréntesis y puntos.
    Si no tiene '+' al inicio, se lo agrega.
    """
    s = clean_cell(value)
    if not s:
        return ""
    
    # quitar caracteres comunes
    s = re.sub(r"[()\s\-\.]", "", s)
    
    # asegurar que empiece con +
    if not s.startswith("+"):
        s = "+" + s
        
    return s

def build_phone(celular, telefono1, telefono2): 
    """
    Formato CloudTalk:
    - numero1:numero2:numero3:numero4 (solo los que existan)
    - Deduplica preservando el orden (primer aparición gana)
    Orden: CELULAR, TELEFONO 1, TELEFONO 2, TELEFONO EMPRESA
    """
    pieces = [
        normalize_phone_piece(celular),
        normalize_phone_piece(telefono1),
        normalize_phone_piece(telefono2),
    ]

    seen = set()
    phones = []
    for p in pieces:
        if not p:
            continue
        if p in seen:
            continue
        seen.add(p)
        phones.append(p)

    return ":".join(phones)


# ==============================
# PAÍS → ISO-3166-1 ALPHA-2 (LITERAL)
# ==============================
ISO2_MAP = {
    # Norte / Centroamérica / Caribe
    "México": "MX",
    "Mexico": "MX",
    "Guatemala": "GT",
    "Belice": "BZ",
    "Honduras": "HN",
    "El Salvador": "SV",
    "Nicaragua": "NI",
    "Costa Rica": "CR",
    "Panamá": "PA",
    "Panama": "PA",
    "Cuba": "CU",
    "Haití": "HT",
    "República Dominicana": "DO",
    "Republica Dominicana": "DO",
    "Puerto Rico": "PR",

    # Sudamérica
    "Colombia": "CO",
    "Venezuela": "VE",
    "Ecuador": "EC",
    "Perú": "PE",
    "Peru": "PE",
    "Bolivia": "BO",
    "Chile": "CL",
    "Argentina": "AR",
    "Uruguay": "UY",
    "Paraguay": "PY",
    "Brasil": "BR",
    "Guyana": "GY",
    "Surinam": "SR",
}

def country_to_iso2(country_value) -> str:
    """
    Convierte PAÍS exactamente como viene (con tildes y mayúsculas)
    a código ISO-3166-1 alpha-2.
    """
    raw = safe_str(country_value)
    if not raw:
        return ""

    iso = ISO2_MAP.get(raw)
    if iso:
        return iso

    # fallback explícito para detectar errores de escritura
    return raw

# ==============================
# EXCEL: FORZAR COLUMNA A TEXTO
# ==============================
def set_excel_column_as_text(filepath: str, header_name: str):
    wb = load_workbook(filepath)
    ws = wb.active

    col_idx = None
    for cell in ws[1]:
        if str(cell.value).strip() == header_name:
            col_idx = cell.column
            break

    if col_idx is None:
        wb.close()
        raise ValueError(f"No se encontró la columna '{header_name}'")

    for row in range(2, ws.max_row + 1):
        c = ws.cell(row=row, column=col_idx)
        c.number_format = "@"
        if c.value is not None:
            c.value = str(c.value)

    wb.save(filepath)
    wb.close()

# ==============================
# MAIN
# ==============================
def main():
    campana_raw = input("Ingrese el nombre de la campaña: ")
    campana = sanitize_filename(campana_raw)

    if not os.path.exists(INPUT_XLSX):
        raise FileNotFoundError(f"No se encontró el archivo: {INPUT_XLSX}")

    df = pd.read_excel(INPUT_XLSX, dtype=str)
    df.columns = [c.strip() for c in df.columns]

    required_columns = [
        "INTERLOCUTOR 1",
        "CELULAR",
        "TELEFONO 1",
        "TELEFONO 2",
        "TELEFONO EMPRESA",
        "CORREO ELECTRÓNICO CORPORATIVO",
        "CORREO ELECTRÓNICO PERSONAL",
        "CARGO",
        "NOMBRE COMERCIAL EMPRESA",
        "CIUDAD / PROVINCIAS",
        "PAÍS",
        "INDUSTRIA",
        "PÁGINA WEB"
    ]

    missing = [c for c in required_columns if c not in df.columns]
    if missing:
        raise ValueError(f"Faltan columnas en el Excel: {missing}")

    output_df = pd.DataFrame({
        "name": df["INTERLOCUTOR 1"].map(clean_cell),

        "phone": [
            build_phone(cel, tel1, tel2)
            for cel, tel1, tel2 in zip(
                df["CELULAR"], df["TELEFONO 1"], df["TELEFONO 2"]
            )
        ],

        "email": df["CORREO ELECTRÓNICO CORPORATIVO"].map(clean_cell),
        "email2": df["CORREO ELECTRÓNICO PERSONAL"].map(clean_cell),
        "title": df["CARGO"].map(clean_cell),
        "company": df["NOMBRE COMERCIAL EMPRESA"].map(clean_cell),

        "address": "",
        "city": df["CIUDAD / PROVINCIAS"].map(clean_cell),
        "state": "",
        "zip": "",

        # ISO-2 literal
        "country": df["PAÍS"].map(country_to_iso2),

        "usc": df["INDUSTRIA"].map(clean_cell),
        "website": df["PÁGINA WEB"].map(clean_cell),

        "campana": clean_cell(campana)
    }, columns=OUTPUT_COLUMNS)

    output_df = output_df.astype(str).replace("nan", "").replace("None", "")

    output_path = os.path.join(OUTPUT_DIR, f"{campana}.xlsx")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        output_df.to_excel(writer, index=False, sheet_name="Sheet1")

    # Forzar columnas críticas a texto
    set_excel_column_as_text(output_path, "phone")
    set_excel_column_as_text(output_path, "country")

    print("\nExcel generado correctamente:")
    print(output_path)
    print(f"Total de filas: {len(output_df)}")

if __name__ == "__main__":
    main()

import openpyxl
import argparse

def main(book_name):
    # Cargar el libro de Excel
    book = openpyxl.load_workbook('./Data/Base global/GlobalDataUpdated-30-3-2026.xlsx')

    # Seleccionar la hoja activa
    sheet = book.active
    string = """
Ale Contratistas SRL.
Girod Peru
ALPHA FINANCE SAC
AGZ Transportes
Almapo S.R.L
SPL Perú
Iolsa
CENADIM - DIGEMID - Ministerio de Salud Perú
Corporación Mayo SAC
Chung & Tong Ingenieros
Pacific Freezing Company
ARCOPA SA (GROUPE ADRIEN)
Villa Salud
Municipalidad de San Luis
GRUPO TRANSPESA OFICIAL
Distribuidora Bajopontina S.A.
SEDALIB SA
Intercorp Financial Services
Diviso Grupo Financiero
Coinsa - Constructores Interamericanos SAC
Agroindustria Santa María SAC
GRUPO LIMA BUS
Gobierno Regional Huanuco
Grupo Econocable
WOW Empresas
LC Perú
AGRICOLA CERRO PRIETO S.A.
Grupo Transmeridian
FARENET
Corporación SIC
Municipalidad Provincial de San Martín
SEAFROST PERÚ
Nubyx Perú
Constructora MPM SA
Clinica San Juan de Dios - Lima
IMECON
Banco Cencosud
Asica Farms
Pesquera Centinela
Oltursa - Transporte y Carga
Mapfre Peru
Petramás S.A.C
Lari Contratistas
Municipalidad de Jesús María
Tecno Fast Perú
Somos ImpulsA365
OTASS Perú
Grupo Vallenorte
Movil Bus
AJANI
Hospital de Emergencias Villa El Salvador
CHINA GEZHOUBA GROUP COMPANY PERÚ
Factotal Perú
Dohwa Engineering Co., Ltd. LATAM
Procesos de Medios de Pago S.A.
CONCYSSA S.A.
CONCREMAX
EOM GRUPO
Municipalidad Provincial de Piura
Peruvian Airlines
CARTAVIO S.A.A.
Inkia Energy
Gobierno Regional Piura
Hospital Regional de Ica
CCECC PERU
Freyssinet Geoquest Perú S.A.C.
Esmeralda Corp SAC
FINANCIERA QAPAQ S.A.
Clínica San Gabriel
TOTAL Servicios Financieros
Coopac KORI
Gobierno Regional de Arequipa
Urbano Perú
Electro Sur Este SAA
YOFC Perú
IMCO Servicios S.A.C.
Caja Cencosud Scotiabank
Olva Courier
Transaltisa S.A.
Clínica Javier Prado
Chimu Agropecuaria S.A.
Gobierno Regional De Ancash Sede Central
SIMA PERÚ
OBRAINSA
Clínica Good Hope
Ovosur
FINANCIERA CREDINKA
Grupo Peru Alfa
WOW Perú
Financiera ProEmpresa
Pamolsa - Carvajal Empaques PE
Huevos La Calera
Equifax Perú
Sinohydro Corporation Limited Sucursal del Perú
SERVICIOS PETROLEROS Y CONSTRUCCIONES SEPCON S.A.C.
Grupo TDM
Menorca Inversiones
Auto Taxi Satelital
Enel Perú
Grupo Palmas
PNSR - Programa Nacional de Saneamiento Rural
Fiberlux
Rintisa
IVC CONTRATISTAS GENERALES S.A.
Ajinomoto del Perú S.A.
Municipalidad de Lima
Lima Expresa
YURA S.A.
Grupo Santa Elena
Machu Picchu Foods
Clínica Anglo Americana
Electro Oriente
SANNA salud
Nextel del Perú S.A.
AUSTRAL GROUP S.A.A.
Gobierno Regional del Callao
Municipalidad de Ventanilla
Provias Descentralizado
SACYR PERÚ
VISIVA
Viru Group
HAUG S.A.
Caja Metropolitana de Lima
Perufarma S.A.
Municipalidad Metropolitana de Lima
Pesquera Diamante
Financiera Oh!
OHLA PERÚ
LOS ANDES
Hospital de la Solidaridad
TRUPAL
INEI- Instituto Nacional de Estadística e Informática
Instituto Nacional de Salud del Niño San Borja
Sharf
Caja Maynas
Compañía Nacional de Chocolates de Perú - Grupo Nutresa
Bancom Perú
Bitel Perú
Clínica Ricardo Palma
PANDERO S.A. EAFC
UNACEM
Clinica San Felipe S.A.
izipay
AFP Integra
Defensoría del Pueblo
Jurado Nacional de Elecciones
Banco Falabella Perú
Industrias San Miguel
Movistar Empresas Perú
Indecopi Oficial
PeruRail
Mota-Engil Perú S.A.
Cobra Perú
Hospital María Auxiliadora
Credicorp
Talma Servicios Aeroportuarios S.A.
Cementos Pacasmayo SAA
Prima AFP
Backus
Municipalidad Distrital de Echarati
Municipalidad de Miraflores
Luz del Sur
Cruz del Sur
San Fernando
CORPAC S.A.
GRUPO DISTRILUZ
Caja Ica
GyM | Grupo Graña y Montero
La Positiva Seguros
Clínica Internacional
BanBif - Banco Interamericano de Finanzas
COSAPI
CONGRESO DE LA REPUBLICA DEL PERU
Banco Central de Reserva del Perú - BCRP
OSINERGMIN
Banco Ripley Perú
Mibanco, banco de la Microempresa
JJC Contratistas Generales S.A.
Ministerio de Cultura del Perú
Alicorp
Registro Nacional de Identificación y Estado Civil - RENIEC
Claro Perú
Banco Pichincha Perú
TASA
Arca Continental Lindley
Caja Arequipa
Rimac Seguros y Reaseguros
BBVA en Perú
Interbank
Banco de Crédito BCP
Pacífico Seguros
EsSalud
"""

    # Lista de empresas
    EMPRESAS = string.split("\n")
    # Limpiar y normalizar los nombres de las empresas
    EMPRESAS_CLEAN = [x.upper().strip().replace("  ", " ").replace(",", "").replace(".", "").replace("(", "").replace(")", "") for x in EMPRESAS]

    # Crear una nueva hoja para los datos filtrados
    new_sheet = book.create_sheet("Filtrados")

    # Buscar la columna "NOMBRE COMERCIAL EMPRESA"
    col_index = None
    for col in sheet.iter_cols(1, sheet.max_column):
        # if col[0].value == "URL LINKEDIN EMPRESA": 
        if col[0].value == "NOMBRE COMERCIAL EMPRESA":
            col_index = col[0].column
            break

    if col_index:
        # Copiar el encabezado a la nueva hoja
        for cell in sheet[1]:
            new_sheet[cell.coordinate] = cell.value

        new_row_index = 2  # Comenzar desde la segunda fila para los datos
        # Iterar sobre las filas y copiar las que cumplen con el criterio
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            cell_value = row[col_index-1].value
            if cell_value and cell_value.upper().strip().replace("  ", " ").replace(",", "").replace(".", "").replace("(", "").replace(")", "") in EMPRESAS_CLEAN:
                
                for cell in row:
                    new_sheet.cell(row=new_row_index, column=cell.column, value=cell.value)
                new_row_index += 1

    # Eliminar la hoja original si lo deseas
    book.remove(sheet)

    # Guardar el archivo modificado
    book.save(f'./Extraídos/{book_name}.xlsx')
    for empresa in EMPRESAS_CLEAN:
        print(empresa)
    print(f"Archivo guardado como '{book_name}.xlsx'")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Filtrar datos de un archivo de Excel y guardarlos en un nuevo archivo.")
    parser.add_argument("book_name", help="Nombre del archivo de salida (sin extensión)")
    args = parser.parse_args()
    main(args.book_name)
import openpyxl
import argparse

def main(book_name, country):
    # Cargar el libro de Excel
    book = openpyxl.load_workbook('./Data/GlobalDataUpdated29-05-2025.xlsx')
    sheet = book.active

    # Crear una nueva hoja para los datos filtrados
    new_sheet = book.create_sheet("Filtrados")

    # Buscar las columnas necesarias
    col_name_index = None
    col_country_index = None
    for col in sheet.iter_cols(1, sheet.max_column):
        header = str(col[0].value).strip().upper() if col[0].value else ""
        if header == "NOMBRE COMERCIAL EMPRESA":
            col_name_index = col[0].column
        if header == "PAÍS":
            col_country_index = col[0].column
    
    if col_name_index and col_country_index:
        # Copiar el encabezado a la nueva hoja
        for cell in sheet[1]:
            new_sheet[cell.coordinate] = cell.value

        new_row_index = 2  # Comenzar desde la segunda fila para los datos
        # Iterar sobre las filas y copiar las que cumplen con el criterio
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            name_value = row[col_name_index-1].value
            country_value = row[col_country_index-1].value
            if name_value and country_value:
                name_clean = str(name_value).upper().strip().replace("  ", " ").replace(",", "").replace(".", "").replace("(", "").replace(")", "")
                country_clean = str(country_value).strip().upper()
                if "UNIVERSIDAD" in name_clean and country_clean == country.strip().upper():
                    for cell in row:
                        new_sheet.cell(row=new_row_index, column=cell.column, value=cell.value)
                    new_row_index += 1

    # Eliminar la hoja original si lo deseas
    book.remove(sheet)

    # Guardar el archivo modificado
    book.save(f'./Nominaciones/{book_name}.xlsx')
    print(f"Archivo guardado como '{book_name}.xlsx'")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Filtrar empresas con 'universidad' en el nombre y país específico.")
    parser.add_argument("book_name", help="Nombre del archivo de salida (sin extensión)")
    parser.add_argument("country", help="Nombre del país a filtrar (ejemplo: México)")
    args = parser.parse_args()
    main(args.book_name, args.country)

